#This module loads the timetable of a given ddepartment for a given year and returns the data stucture.
#It also generates venues timetable on reading timetable cell contents and stores the returned structure
#
##
#
#

from openpyxl import load_workbook
import os
from openpyxl.utils.cell import column_index_from_string
import pickle

# import venue_generator

class Timetables:

    def __init__(self):

        self.lecture = {}
        self.timetable = {}

    def _time_structure(self, sheet, starting_point):

        self.time_format = {}

        for item in range(2, sheet.max_column + 1):
            _time_value = sheet.cell(row=starting_point, column=item).value
            if _time_value is None:
                return self.time_format
            else:
                self.time_format[_time_value] = self.lecture
        else:
            return self.time_format

    def _event_time(self, time_format, key_pos):

        total = 0

        for key, value in time_format.items():
            total = total + 1
            if total == key_pos:
                return key
            else:
                continue

        else:
            pass

    def unit_group_generator(self, cell_value, institution_name):

        self.timetable = {}
        for item in range(cell_value.count(',') + 1):

            if cell_value.find(',') == -1:
                value = cell_value
            else:
                value = cell_value[:cell_value.find(',')]
                cell_value = cell_value[cell_value.find(',') + 1:]

            if len(value) == 6:
                unit_name = value
            else:
                if len(value) == 4:
                    if value[:2].upper() == 'GP':
                        group_name = value
                        unit_name = unit_name + ',' + group_name
                    else:
                        venue_name = value
                        self.timetable[unit_name] = venue_name
                        # assert venue_generator._venue_generator_section(institution_name, venue_name, day, time, unit_name)
                else:
                    self.timetable[unit_name] = value
                    # assert venue_generator._venue_generator_section(institution_name, venue_name, day, time, unit_name)
        else:
            return self.timetable


con = Timetables()


def _timetable_section(institution_name, school_name, cell_value, starting_point, dept_name, dept_year, semester):
    timetable_structure = {}
    _time_format = {}

    def tt_writter(timetable_structure):

        filename = open(os.path.normpath(
                                os.getcwd() + os.sep + os.pardir) + '\\' + 'data' + '\\' + institution_name + '\\' + school_name + '\\' + dept_name + '\\' + dept_year + dept_name + '_' + semester + '_timetable.TT', 'wb' )
        pickle.dump(timetable_structure, filename)
        filename.close()

    try:
        wb = load_workbook(os.path.normpath(
            os.getcwd() + os.sep + os.pardir) + '\\' + 'data' + '\\' + institution_name + '\\' + school_name + '_structure.xlsx')
        sheet = wb['timetables']

        # the if statement eliminates the bug created by the for loop in line 97.
        if starting_point >= sheet.max_row:
            return timetable_structure
        else:

            try:

                filelist = os.listdir(os.path.normpath(os.getcwd() + os.sep + os.pardir) + '\\' + institution_name +
                                      '\\' + school_name + '\\' + dept_name + '\\' + dept_year)

            except FileNotFoundError:
                try:
                    os.mkdir(os.path.normpath(os.getcwd() + os.sep + os.pardir) + '\\' + institution_name +
                             '\\' + school_name + '\\' + dept_name + '\\' + dept_year)
                    filelist = []
                except FileExistsError:
                    pass

            if (dept_name + '_' + semester + '_timetable.TT') in filelist:
                pass
            else:
                for item in range(starting_point + 1, sheet.max_row + 1):

                    Day_name = sheet.cell(row=item, column=1).value
                    _time_format = con._time_structure(sheet, starting_point)
                    timetable_structure[Day_name] = _time_format

                    for item1 in range(2, sheet.max_column + 1):
                        cell_value1 = sheet.cell(row=item, column=item1).value
                        if cell_value1 is None:
                            continue
                        else:
                            event_time = con._event_time(_time_format, column_index_from_string(
                                sheet.cell(row=item, column=item1).coordinate[:1]) - 1)
                            _time_format[event_time] = con.unit_group_generator(cell_value1, institution_name)
                            timetable_structure[Day_name] = _time_format
                    else:
                        pass

                    if len(timetable_structure) == cell_value:
                        try:

                            os.mkdir(os.path.normpath(
                                os.getcwd() + os.sep + os.pardir) + '\\' + 'data' + '\\' + institution_name + '\\' + school_name + '\\' + dept_name + '\\' + dept_year)
                            try:
                                assert os.path.exists(os.path.normpath(
                                os.getcwd() + os.sep + os.pardir) + '\\' + 'data' + '\\' + institution_name + '\\' + school_name + '\\' + dept_name + '\\' + dept_year)
                            except AssertionError:
                                # write in the error log.
                                pass

                            tt_writter(timetable_structure)

                        except FileExistsError:
                            con.timetable_writter(institution_name, school_name, dept_name, dept_year, timetable_structure)
                            tt_writter(timetable_structure)

                        return timetable_structure
                    else:
                        continue
                else:
                    pass

    except ModuleNotFoundError:
        pass
    except FileNotFoundError:
        pass

    return timetable_structure
